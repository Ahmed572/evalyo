'''
import openpyxl

wb = openpyxl.load_workbook('test.xlsx')

type(wb)


print (wb.get_sheet_names())
 
sheet = wb.sheet_by_index(0)
sheet.cell_value(0,0)

'''

import openpyxl 
import os
wb = openpyxl.load_workbook('courses-list-2017-02-28_19.11.01.xlsx')

sheet = wb.get_sheet_by_name('Worksheet')

#print(sheet['B1'].value)
#print(sheet.cell(row=2,column=2))
#print(sheet.cell(row=2,column=2).value)

maxrow = sheet.max_row
maxcolumn = sheet.max_column
#print(maxrow)
#print(maxcolumn)
'''
for i in range(1,maxcolumn):
	print(sheet.cell(row=1,column=i).value)
'''	
'''
1 Catalog OID no 
2 Catalog Name 
3 School/College Code
4 School/College Name
5 Deptarment Code
6 Department Name
7 Entity Type
8 Entity OID
9 Entity Name
10 Course OID
11 Course Type
12 Prefix
13 Code
14 Name 
15 Huntr Core
16 Pluralism and Diversity 
17 GER
18 Description 
19 Description(Rendered)
20 Description(Rendered no HTML)
21 Notes
22 Notes(Rendered)
23 Notes(Rendered no HTML)
24 cross-listed
25 prereq:
26 coreq:
27 prereq or coreq:
28 Hours
29 Credits
30 When offered
31 Is Active
32 Creation Date
33 Last Modified 
34 Program Usage
'''

'''
for i in range(1,maxrow):
	for j in range(1,maxcolumn):
		print(sheet.cell(row=i,column=j).value) 
'''
Catalog_OID = []
Catalog_Name =[]
School_College_Code =[]
School_College_Name =[]
Deptarment_Code =[]
Department_Name =[]
Entity_Type =[]
Entity_OID =[]
Entity_Name =[]
Course_OID =[]
Course_Type =[]
Prefix =[]
Code =[]
Name =[]
Hunter_Core =[]
Pluralism_and_Diversity =[]
GER =[]
Description =[]
Description_Rendered =[]
Description_Rendered_no_HTML =[]
Notes =[]
Notes_Rendered =[]
Notes_Rendered_no_HTML =[]
cross_listed =[]
prereq =[]
coreq =[]
prereq_or_coreq =[]
Hours =[]
Credits =[]
When_offered =[]
Is_Active =[]
Creation_Date =[]
Last_Modified =[]
Program_Usage =[]
for i in range(1,maxcolumn):
	for j in range(2,maxrow):
		if i == 1:
			Catalog_OID.append(sheet.cell(row=j,column=i).value)
		if i == 2:
			Catalog_Name.append(sheet.cell(row=j,column=i).value)
		if i == 3:
			School_College_Code.append(sheet.cell(row=j,column=i).value)
		if i == 4:
			School_College_Name.append(sheet.cell(row=j,column=i).value)
		if i == 5:
			Deptarment_Code.append(sheet.cell(row=j,column=i).value)
            
		if i == 6:
			Department_Name.append(sheet.cell(row=j,column=i).value)

		if i == 7:
			Entity_Type.append(sheet.cell(row=j,column=i).value)

		if i == 8:
			Entity_OID.append(sheet.cell(row=j,column=i).value)

		if i == 9:
			Entity_Name.append(sheet.cell(row=j,column=i).value)

		if i == 10:
			Course_OID.append(sheet.cell(row=j,column=i).value)

		if i == 11:
			Course_Type.append(sheet.cell(row=j,column=i).value)

		if i == 12:
			Prefix.append(sheet.cell(row=j,column=i).value)

		if i == 13:
			Code.append(sheet.cell(row=j,column=i).value)

		if i == 14:
			Name.append(sheet.cell(row=j,column=i).value)

		if i == 15:
			Hunter_Core.append(sheet.cell(row=j,column=i).value)

		if i == 16:
			Pluralism_and_Diversity.append(sheet.cell(row=j,column=i).value)

		if i == 17:
			GER.append(sheet.cell(row=j,column=i).value)

		if i == 18:
			Description.append(sheet.cell(row=j,column=i).value)

		if i == 19:
			Description_Rendered.append(sheet.cell(row=j,column=i).value)

		if i == 20:
			Description_Rendered_no_HTML.append(sheet.cell(row=j,column=i).value)

		if i == 21:
			Notes.append(sheet.cell(row=j,column=i).value)

		if i == 22:
			Notes_Rendered.append(sheet.cell(row=j,column=i).value)

		if i == 23:
			Notes_Rendered_no_HTML.append(sheet.cell(row=j,column=i).value)

		if i == 24:
			cross_listed.append(sheet.cell(row=j,column=i).value)

		if i == 25:
			prereq.append(sheet.cell(row=j,column=i).value)

		if i == 26:
			coreq.append(sheet.cell(row=j,column=i).value)

		if i == 27:
			prereq_or_coreq.append(sheet.cell(row=j,column=i).value)

		if i == 28:
			Hours.append(sheet.cell(row=j,column=i).value)

		if i == 29:
			Credits.append(sheet.cell(row=j,column=i).value)

		if i == 30:
			When_offered.append(sheet.cell(row=j,column=i).value)

		if i == 31:
			Is_Active.append(sheet.cell(row=j,column=i).value)

		if i == 32:
			Creation_Date.append(sheet.cell(row=j,column=i).value)

		if i == 33:
			Last_Modified.append(sheet.cell(row=j,column=i).value)

		if i == 34:
			Program_Usage.append(sheet.cell(row=j,column=i).value)






#print(Catalog_OID)
#print(Catalog_Name)
#print(School_College_Code)
#print(School_College_Name)
#print(Deptarment_Code)
#print(Department_Name)
#print(Entity_Type)
#print(Entity_OID)
#print(Entity_Name)
#print(Course_OID)
#print(Course_Type)
#print(Prefix)
#print(Code)
#print(Name)
#print(Hunter_Core) 
#print(Pluralism_and_Diversity)
#print(GER) 
#print(Description) 
#print(Description_Rendered) 
#print(Description_Rendered_no_HTML) 
#print(Notes)
#print(Notes_Rendered)
#print(Notes_Rendered_no_HTML)
#print(cross_listed)
#print(prereq) 
#print(coreq) 
#print(prereq_or_coreq)
#print(Hours) 
#print(Credits) 
#print(When_offered) 
#print(Is_Active) 
#print(Creation_Date) 
#print(Last_Modified)
#print(Program_Usage)


#print("__")
#print(sheet.max_row)
#print(sheet.max_column)