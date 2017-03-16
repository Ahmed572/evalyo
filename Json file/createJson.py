import openpyxl 
import os
import json

wb = openpyxl.load_workbook('courses-list-2017-02-28_19.11.01.xlsx')

sheet = wb.get_sheet_by_name('Worksheet')

maxrow = sheet.max_row
maxcolumn = sheet.max_column

'''
catalog_Name #2
Dept_Name #6
Entity_Name #9
Course_Type #11
Name #14
Hours #28
Credits #29
Is_Active #31
'''



book = []
for i in range(2,3):
	for j in range(1,maxcolumn):
		if j == 2 or j == 6 or j == 9 or j == 11 or j == 14 or j == 28 or j ==29 or j == 31:
			book.append(sheet.cell(row=i,column=j).value)

s = json.dumps(book,indent = 4, sort_keys=True)

with open("book.json","w") as f:
	f.write(s)



